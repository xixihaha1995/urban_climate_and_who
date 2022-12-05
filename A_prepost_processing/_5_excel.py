# Import the openpyxl library.
from openpyxl import Workbook

# Create a new Excel workbook.
wb = Workbook()

# Get the active sheet.
sheet = wb.active

# Generate 6 random numbers and add them to the first column of the Excel sheet.
import random
for i in range(1, 7):
  sheet.cell(row=i, column=1).value = random.randint(1, 100)

# Add the formulas to calculate the squares of the numbers in the first column and
# add them to the second column of the Excel sheet.
for i in range(1, 7):
  sheet.cell(row=i, column=2).value = '=A{}*A{}'.format(i, i)

# Save the Excel file.
wb.save('random_numbers.xlsx')
